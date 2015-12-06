#!/usr/bin/env python
# PYTHON_ARGCOMPLETE_OK
from flask.ext.script import Manager, Command, Option
from sidr import app


class ImportOldHira(Command):
    "import old hira database"

    def get_options(self):
        return [
            Option('--name', dest='name', required=True),
            Option('--name-display', dest='name_display', required=True),
            Option('--restrict-countries', dest='restrict_countries', required=True),
            Option('--db', dest='db', required=True)
        ]

    def run(self, name, name_display, restrict_countries, db):
        import pymysql
        from sidr import orm, models

        data = {
            'name': name,
            'name_display': name_display,
            'restrict_countries': restrict_countries.split(',')
        }

        domain = models.Domain.af_save(models.User(), data)
        conn = pymysql.connect(host='localhost', user='root', db=db, charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)

        user_map = self.import_users(domain, conn)
        self.import_leads(domain, conn, user_map)

    def import_users(self, domain, conn):
        from sidr import orm, models
        from sidr.models.model import random_string

        user_map = {}

        with conn.cursor() as cursor:
            cursor.execute('SELECT * FROM user')
            for row in cursor:
                user = models.User.find_first(email=row['email'])
                if user is None:
                    data = {
                        'email': row['email'],
                        'orgnization': row['org'],
                        'password': random_string()
                    }
                    user = models.User(**data)
                    user.save()

                user.connect_domain(domain['id'])
                user_map[row['id']] = user.id

        return user_map

    def import_leads(self, domain, conn, user_map):
        from sidr import orm, models

        with conn.cursor() as cursor:
            cursor.execute('SELECT * FROM source')
            for row in cursor:
                data = {
                    'lead_type': 'url',
                    'name': '',
                    'url': row['link'],
                    'website': row['site'],
                    'published_at': row['date'],
                    'information_at': row['createdAt'],
                    'domain_id': domain['id']
                }
                models.Lead.af_save(models.User.find_first(id=user_map[row['user_id']]), data)


class ImportTags(Command):
    "import tags from xlsx"

    """
    'affected': {
        'worksheet': 'Affected group',
        'columns': {

        }
    },
    """

    import_map = {
        'event_type': {
            'worksheet': 'Event Type',
            'columns': {
                'A': 'name',
                'B': 'title',
                'C': 'parent_id',
                'D': 'description'
            }
        },
        'sector': {
            'worksheet': 'Sector and sub sector',
            'columns': {
                'A': 'name',
                'B': 'title',
                'C': 'parent_id',
                'D': 'description'
            }
        },
        'source': {
            'worksheet': 'Source',
            'columns': {
                'A': 'name',
                'B': 'title'
            }
        },
        'status': {
            'worksheet': 'Status',
            'columns': {
                'A': 'name',
                'B': 'title'
            }
        },
        'timeline': {
            'worksheet': 'Problem Timeline',
            'columns': {
                'A': 'name',
                'B': 'title'
            }
        },
        'underlying': {
            'worksheet': 'Underlying Factor',
            'columns': {
                'A': 'name',
                'B': 'title',
                'C': 'parent_id',
                'D': 'description'
            }
        },
        'affected': {
            'worksheet': 'Affected group',
            'columns': {
                'A': 'name',
                'B': 'title',
                'C': 'parent_id',
                'D': 'description'
            }
        },
        'vulnerable': {
            'worksheet': 'Vulnerable group',
            'columns': {
                'A': 'name',
                'B': 'title',
                'C': 'parent_id',
                'D': 'description'
            }
        },
    }

    def get_options(self):
        return [
            Option('-n', '--filename', dest='filename'),
        ]

    def process_ws(self, tag_class, ws, ws_map):
        from sidr import models, validator
        rows = ws.iter_rows()
        next(rows)
        for row in rows:
            for cell in row:
                print(str(cell.column) + " " + str(cell.value))
            data = {ws_map['columns'][cell.column]: cell.value for cell in row if cell.column in ws_map['columns']}

            if data['name'] is None:
                continue

            data['tag_class'] = tag_class['name']
            current = models.Tag.find_first({'name': data['name'], 'tag_class': data['tag_class']})
            obj_id = current.id if current is not None else None
            if 'parent_id' in data:
                if 'parent_id' is None:
                    del data['parent_id']
                else:
                    parent = models.Tag.find_first({'name': data['parent_id'], 'tag_class': data['tag_class']})
                    if parent is None:
                        raise Exception('parent set but not parent in db %s' % data['parent_id'])
                    data['parent_id'] = parent.id

            try:
                models.Tag.af_save(models.User(), data, obj_id=obj_id)
            except validator.ApiValidationError as e:
                raise Exception(repr(data) + ' IN ' + repr(e))

    def run(self, filename):
        from openpyxl import load_workbook
        from sidr import orm, models
        from sidr.orm import db


        wb = load_workbook(filename, use_iterators=True)
        worksheet_names = wb.get_sheet_names()
        for tag_class in models.Tag.af_tag_classes()['result']:
            if tag_class['name'] in self.import_map:
                ws_map = self.import_map[tag_class['name']]
                ws_name = ws_map['worksheet']
                if ws_name in worksheet_names:
                    ws = wb.get_sheet_by_name(name=ws_name)
                    self.process_ws(tag_class, ws, ws_map)

        db.session.commit()


class ParseGeonames(Command):
    "import geonames"

    def build_db(self):
        import csv
        import sys
        from sidr import orm, models
        from sidr.orm import db
        csv.field_size_limit(sys.maxsize)

        pmap = {}
        gmap = {}
        umap = {}
        with open('import/hierarchy.txt', 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter='\t')
            for row in reader:
                pmap[row[1]] = row[0]
                umap[row[0]] = True
                umap[row[1]] = True
                # print(repr(row))
        with open('import/allCountries.txt', 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter='\t')
            for row in reader:
                if row[0] in umap:
                    uid = int(row[0])
                    obj = {}
                    if str(uid) in pmap:
                        obj['parent_id'] = pmap[str(uid)]
                    obj['id'] = uid
                    obj['name'] = row[1]
                    obj['asciiname'] = row[2]
                    if len(row[3]) > 0:
                        obj['name_alternate'] = [x.strip() for x in row[3].split(',')]
                    else:
                        obj['name_alternate'] = []
                    obj['latitude'] = row[4]
                    obj['longitude'] = row[5]
                    obj['feature_code'] = row[7]
                    obj['country_code'] = row[8]
                    geoname = models.Geoname(**obj)
                    geoname.save()
                    db.session.commit()

    def indexloc(self):
        from sidr.orm import db
        import json
        from whoosh.index import create_in, open_dir, exists_in
        from whoosh import fields, qparser, query
        schema = fields.Schema(gid=fields.TEXT(stored=True), country_code=fields.ID(stored=True), names=fields.NGRAMWORDS(stored=True, minsize=3, maxsize=15))
        if not exists_in("indexer", indexname="adms"):
            ix = create_in("indexer", schema, indexname="adms")
        ix = open_dir("indexer", indexname="adms")
        writer = ix.writer()

        """
        with ix.searcher() as s:
            qp = qparser.QueryParser("names", schema=ix.schema)
            q = qp.parse(u"Westonia")
            # results = s.search(q, limit=20, filter=query.Term("country_code", "AU"))
            results = s.documents()
            # results = searcher.search('hey', terms=True)
            # qp = qparser.QueryParser("content", ix.schema)
            # results = searcher.search(user_q)
            for res in results:
                print(repr(res))
        """

        rows = db.engine.execute('SELECT * FROM geoname')
        for row in rows:
            writer.add_document(gid=str(row['id']), country_code=row['country_code'], names="%s , %s , %s" % (row['name'], row['asciiname'], row['name_alternate']))
        writer.commit()

    def run(self):
        # self.build_db()
        self.indexloc()


manager = Manager(app)
manager.add_command('import_tags', ImportTags())
manager.add_command('import_db', ImportOldHira())
manager.add_command('parse_geonames', ParseGeonames())

if __name__ == "__main__":
    manager.run()
